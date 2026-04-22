#!/usr/bin/env python3
"""
P13.3 / T1: Ingest service-page metadata from xlsx as semantic-enrichment docs.

Source: RAG_DATA/страницы - родители товаров.xlsx (80 rows, 8 cols).
Schema: ID | Meta Title | Meta Description | H1 | Название | Полное название |
        Вводный текст | Описание

Each row → one doc with doc_type="service_page". The `page_id` matches
PARENT_SECTION in offers.csv / goods.csv, so retriever can join semantically.

xlsx has NO slug/URL column (verified) — service URLs continue to use
existing `/product/{ID}` scheme via ingest.py:242. This file only adds
SEO-quality natural-language enrichment to the dense index, which closes
gaps like chat #96/300 ("стенды для школы" — semantic miss).

Output: data/service_page_docs.jsonl
"""
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import click

PROJECT_ROOT = Path(__file__).parent.parent
RAG_DATA = PROJECT_ROOT.parent / "RAG_DATA"
OUTPUT_DIR = PROJECT_ROOT / "data"

XLSX_PATH = RAG_DATA / "страницы - родители товаров.xlsx"
OUTPUT_PATH = OUTPUT_DIR / "service_page_docs.jsonl"

GENERATED_AT = datetime.now(timezone.utc).isoformat()


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_searchable(meta_title: str, meta_desc: str, full_title: str,
                     intro: str, description: str) -> str:
    """Concatenate human-readable fields for dense retrieval. Order matters:
    titles first (high signal for keyword match), then descriptions.
    """
    parts = []
    if full_title:
        parts.append(full_title)
    if meta_title and meta_title != full_title:
        parts.append(meta_title)
    if meta_desc:
        parts.append(meta_desc)
    if intro:
        parts.append(intro)
    if description:
        parts.append(description)
    return "\n\n".join(parts)


@click.command()
@click.option("--verbose", is_flag=True, default=False)
def main(verbose: bool):
    if not XLSX_PATH.exists():
        print(f"[ERR] xlsx not found: {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[ERR] openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb.active

    docs = []
    skipped = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        page_id = _norm(row[0])
        meta_title = _norm(row[1])
        meta_desc = _norm(row[2])
        # row[3] is H1 — always null in source, kept for forward-compat
        h1 = _norm(row[3])
        name = _norm(row[4])
        full_title = _norm(row[5])
        intro_text = _norm(row[6])
        description = _norm(row[7])

        if not page_id:
            skipped += 1
            continue
        # Need at least title + (intro or description) to be useful
        if not (full_title or name) or not (intro_text or description or meta_desc):
            skipped += 1
            if verbose:
                print(f"  [SKIP row {row_idx}] page_id={page_id}: insufficient text")
            continue

        searchable = build_searchable(meta_title, meta_desc,
                                      full_title or name,
                                      intro_text, description)

        doc_id = f"service_page_{int(page_id):05d}"
        display_name = full_title or name or meta_title

        doc = {
            "doc_id": doc_id,
            "doc_type": "service_page",
            "searchable_text": searchable,
            "payload": {
                "doc_id": doc_id,
                "doc_type": "service_page",
                "page_id": page_id,
                "parent_section_id": page_id,  # for join with offers.csv PARENT_SECTION
                "name": name,
                "full_title": full_title,
                "meta_title": meta_title,
                "meta_description": meta_desc,
                "h1": h1,
                "intro_text": intro_text,
                "description": description,
                "searchable_text": searchable,
                "display_label": display_name,
            },
            "provenance": {
                "source": "страницы - родители товаров.xlsx",
                "row": row_idx,
                "generated_at": GENERATED_AT,
            },
        }
        docs.append(doc)
        if verbose:
            print(f"  [OK row {row_idx}] {doc_id} {display_name[:60]}")

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        for d in docs:
            f.write(json.dumps(d, ensure_ascii=False) + "\n")

    print(f"\nWritten {len(docs)} service_page docs to {OUTPUT_PATH}")
    print(f"Skipped: {skipped}")


if __name__ == "__main__":
    main()
